#!/usr/bin/env python3
"""Audit Google Ads — termes de recherche -> xlsx multi-onglets.

Usage:
    python analyze.py rapport.csv --objective objectif.txt -o audit.xlsx

Pré-requis:
    pip install -r requirements.txt
    export ANTHROPIC_API_KEY=sk-ant-...
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from anthropic import Anthropic, APIError, RateLimitError
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MODEL = "claude-opus-4-7"
BATCH_SIZE = 40
MAX_RETRIES = 3

COLUMN_ALIASES = {
    # Anglais
    "search term": "term",
    "match type": "match_type",
    "added/excluded": "status",
    "keyword added/excluded": "status",
    "campaign": "campaign",
    "ad group": "ad_group",
    "impr.": "impressions",
    "impressions": "impressions",
    "clicks": "clicks",
    "cost": "cost",
    "conversions": "conversions",
    "conv. value": "conv_value",
    "conv value": "conv_value",
    "cost / conv.": "cpa",
    "cost/conv.": "cpa",
    "ctr": "ctr",
    "avg. cpc": "avg_cpc",
    "average cpc": "avg_cpc",
    # Français
    "terme de recherche": "term",
    "type de correspondance": "match_type",
    "ajouté/exclu": "status",
    "mot clé ajouté/exclu": "status",
    "mot-clé ajouté/exclu": "status",
    "état": "status",
    "campagne": "campaign",
    "groupe d'annonces": "ad_group",
    "clics": "clicks",
    "coût": "cost",
    "valeur de conversion": "conv_value",
    "valeur conv.": "conv_value",
    "coût / conv.": "cpa",
    "coût/conv.": "cpa",
    "cpc moy.": "avg_cpc",
    "taux de conv.": "conv_rate",
}

SYSTEM_PROMPT = """Tu es un expert Google Ads qui audite un rapport de termes de recherche pour identifier les candidats à exclure en mots-clés négatifs.

Pour chaque terme de recherche, tu dois décider:
- "keep": terme aligné avec l'objectif de la campagne, qui convertit ou pourrait convertir
- "exclude": terme hors-sujet, attire une mauvaise audience, gaspille du budget (à ajouter en mot-clé négatif)
- "review": terme ambigu, nécessite une revue manuelle avant décision

Critères d'exclusion typiques (à adapter à l'objectif client):
- Termes informationnels quand l'objectif est transactionnel ("comment", "qu'est-ce que", "définition", "tuto")
- Termes "gratuit", "DIY", "apprendre" quand on vend un service premium
- Noms de concurrents (stratégie parfois volontaire — marquer "review" si doute)
- Recherches d'emploi ("job", "recrutement", "emploi", "salaire", "stage")
- Termes géographiques hors zone cible
- Termes off-topic n'ayant aucun rapport avec l'objectif

Tu dois aussi:
- Repérer les marques concurrentes (competitor_brand = nom exact de la marque, ou null)
- Classer par thème sémantique (ex: "informational", "prix", "concurrent", "emploi", "intent-commercial", "formation", "avis", "off-topic")
- Pour chaque exclusion, suggérer le type de match négatif: "exact" (bloquer uniquement ce terme précis) ou "phrase" (bloquer toute requête contenant cette expression); null si keep/review

Sois rigoureux mais pas trop strict — une mauvaise exclusion fait perdre des clients potentiels. Privilégie "review" en cas de doute."""

CLASSIFICATION_SCHEMA = {
    "type": "object",
    "properties": {
        "classifications": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "term": {"type": "string"},
                    "decision": {"type": "string", "enum": ["keep", "exclude", "review"]},
                    "reason": {"type": "string"},
                    "is_competitor": {"type": "boolean"},
                    "competitor_brand": {"type": ["string", "null"]},
                    "theme": {"type": "string"},
                    "suggested_match_type": {"type": ["string", "null"], "enum": ["exact", "phrase", None]},
                },
                "required": ["term", "decision", "reason", "is_competitor", "competitor_brand", "theme", "suggested_match_type"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["classifications"],
    "additionalProperties": False,
}


def find_header_row(path: Path, max_scan: int = 15) -> int:
    """Retourne l'index de la ligne d'en-tête (Google Ads ajoute souvent 1-3 lignes de métadonnées)."""
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        for i, line in enumerate(f):
            if i >= max_scan:
                break
            low = line.lower()
            if "terme de recherche" in low or "search term" in low:
                return i
    return 0


def parse_number(val) -> float:
    if pd.isna(val) or val in ("", "--", "—"):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s and "." in s:
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def load_csv(path: Path) -> pd.DataFrame:
    header_row = find_header_row(path)
    try:
        df = pd.read_csv(path, sep=None, engine="python", skiprows=header_row, encoding="utf-8-sig")
    except Exception as e:
        raise SystemExit(f"Impossible de lire le CSV ({path}): {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    rename_map = {c: COLUMN_ALIASES[c] for c in df.columns if c in COLUMN_ALIASES}
    df = df.rename(columns=rename_map)

    if "term" not in df.columns:
        raise SystemExit(
            "Aucune colonne 'Terme de recherche' / 'Search term' trouvée. "
            f"Colonnes détectées: {list(df.columns)}"
        )

    df = df[df["term"].notna()]
    df = df[~df["term"].astype(str).str.strip().str.lower().str.startswith(("total", "totaux"))]
    df["term"] = df["term"].astype(str).str.strip()
    df = df[df["term"] != ""]

    for col in ("impressions", "clicks", "cost", "conversions", "conv_value", "cpa", "avg_cpc"):
        if col in df.columns:
            df[col] = df[col].apply(parse_number)

    for col in ("impressions", "clicks", "cost", "conversions", "conv_value"):
        if col not in df.columns:
            df[col] = 0.0

    for col in ("match_type", "status", "campaign", "ad_group"):
        if col not in df.columns:
            df[col] = ""

    df = df.drop_duplicates(subset=["term"], keep="first").reset_index(drop=True)
    return df


def classify_batch(client: Anthropic, objective: str, terms_batch: list[dict], model: str) -> list[dict]:
    user_content = (
        f"**Objectif de la campagne:**\n{objective}\n\n"
        f"**Termes à classifier** ({len(terms_batch)} termes, avec coût / clics / conversions pour contexte):\n"
        f"{json.dumps(terms_batch, ensure_ascii=False, indent=2)}"
    )

    last_err: Exception | None = None
    for attempt in range(MAX_RETRIES):
        try:
            response = client.messages.create(
                model=model,
                max_tokens=16000,
                system=[{"type": "text", "text": SYSTEM_PROMPT, "cache_control": {"type": "ephemeral"}}],
                messages=[{"role": "user", "content": user_content}],
                thinking={"type": "adaptive"},
                output_config={"format": {"type": "json_schema", "schema": CLASSIFICATION_SCHEMA}},
            )
            for block in response.content:
                if block.type == "text":
                    data = json.loads(block.text)
                    return data["classifications"]
            raise RuntimeError("Réponse sans bloc texte")
        except (RateLimitError, APIError) as e:
            last_err = e
            wait = 2 ** attempt
            print(f"  ⚠ Erreur API (tentative {attempt + 1}/{MAX_RETRIES}): {e}. Retry dans {wait}s...", file=sys.stderr)
            time.sleep(wait)
    raise RuntimeError(f"Échec après {MAX_RETRIES} tentatives: {last_err}")


def classify_all(client: Anthropic, objective: str, df: pd.DataFrame, model: str, batch_size: int) -> pd.DataFrame:
    payload = df[["term", "cost", "clicks", "conversions"]].to_dict(orient="records")
    for row in payload:
        row["cost"] = round(row["cost"], 2)
        row["clicks"] = int(row["clicks"])
        row["conversions"] = round(row["conversions"], 2)

    results: dict[str, dict] = {}
    total_batches = (len(payload) + batch_size - 1) // batch_size
    for i in range(0, len(payload), batch_size):
        batch_num = i // batch_size + 1
        batch = payload[i:i + batch_size]
        print(f"  Batch {batch_num}/{total_batches} ({len(batch)} termes)...", flush=True)
        try:
            classifications = classify_batch(client, objective, batch, model)
        except Exception as e:
            print(f"  ✗ Batch {batch_num} échoué: {e}. Termes marqués 'review'.", file=sys.stderr)
            classifications = [
                {
                    "term": r["term"], "decision": "review",
                    "reason": "Classification échouée — à revoir manuellement",
                    "is_competitor": False, "competitor_brand": None,
                    "theme": "unclassified", "suggested_match_type": None,
                } for r in batch
            ]
        for c in classifications:
            results[c["term"]] = c

    classif_df = pd.DataFrame([results.get(t, {
        "term": t, "decision": "review", "reason": "Non classé",
        "is_competitor": False, "competitor_brand": None,
        "theme": "unclassified", "suggested_match_type": None,
    }) for t in df["term"]])

    merged = df.merge(classif_df, on="term", how="left")
    return merged


def fmt_eur(val) -> str:
    try:
        return f"{float(val):,.2f} €".replace(",", " ").replace(".", ",")
    except (ValueError, TypeError):
        return str(val)


def autosize_columns(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col_cells), default=10)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def style_header(ws, n_cols: int) -> None:
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def color_decision(ws, col_idx: int, n_rows: int) -> None:
    colors = {
        "keep": "D1FAE5", "exclude": "FEE2E2", "review": "FEF3C7",
    }
    for row in range(2, n_rows + 2):
        cell = ws.cell(row=row, column=col_idx)
        val = str(cell.value or "").lower()
        if val in colors:
            cell.fill = PatternFill(start_color=colors[val], end_color=colors[val], fill_type="solid")


def build_synthesis(df: pd.DataFrame) -> pd.DataFrame:
    total_cost = float(df["cost"].sum())
    total_conv = float(df["conversions"].sum())
    total_conv_value = float(df["conv_value"].sum())
    wasted = float(df.loc[df["decision"] == "exclude", "cost"].sum())
    top_savings = (
        df[df["decision"] == "exclude"]
        .nlargest(10, "cost")[["term", "cost"]]
        .to_dict(orient="records")
    )
    rows = [
        ("Nombre total de termes", len(df)),
        ("Dépense totale", fmt_eur(total_cost)),
        ("Conversions totales", round(total_conv, 2)),
        ("Valeur de conversion totale", fmt_eur(total_conv_value)),
        ("CPA moyen", fmt_eur(total_cost / total_conv if total_conv else 0)),
        ("ROAS", f"{(total_conv_value / total_cost):.2f}" if total_cost else "—"),
        ("", ""),
        ("Termes à exclure (candidats négatifs)", int((df["decision"] == "exclude").sum())),
        ("Termes à revoir manuellement", int((df["decision"] == "review").sum())),
        ("Termes à garder", int((df["decision"] == "keep").sum())),
        ("", ""),
        ("Budget gaspillé (termes à exclure)", fmt_eur(wasted)),
        ("% du budget gaspillé", f"{(wasted / total_cost * 100):.1f} %" if total_cost else "—"),
        ("", ""),
        ("— TOP 10 ÉCONOMIES POTENTIELLES —", ""),
    ]
    for item in top_savings:
        rows.append((item["term"], fmt_eur(item["cost"])))
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def build_top_spenders(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["term", "decision", "cost", "clicks", "conversions", "conv_value",
            "reason", "theme", "campaign", "ad_group"]
    cols = [c for c in cols if c in df.columns]
    out = df[cols].sort_values("cost", ascending=False).copy()
    out.columns = [
        {"term": "Terme", "decision": "Décision", "cost": "Coût (€)", "clicks": "Clics",
         "conversions": "Conversions", "conv_value": "Valeur conv. (€)",
         "reason": "Justification", "theme": "Thème", "campaign": "Campagne",
         "ad_group": "Groupe d'annonces"}.get(c, c) for c in out.columns
    ]
    return out


def build_candidates(df: pd.DataFrame) -> pd.DataFrame:
    sub = df[df["decision"] == "exclude"].copy()
    sub = sub.sort_values("cost", ascending=False)
    out = sub[["term", "cost", "clicks", "conversions", "suggested_match_type", "reason", "theme"]].copy()
    out.columns = ["Terme à exclure", "Coût gaspillé (€)", "Clics", "Conv.",
                   "Type de match négatif", "Raison", "Thème"]
    return out


def build_competitors(df: pd.DataFrame) -> pd.DataFrame:
    sub = df[df["is_competitor"] & df["competitor_brand"].notna()].copy()
    if sub.empty:
        return pd.DataFrame([{"Marque concurrente": "Aucun concurrent détecté", "Dépense (€)": 0,
                              "Nb termes": 0, "Clics": 0, "Conversions": 0}])
    agg = sub.groupby("competitor_brand").agg(
        depense=("cost", "sum"),
        nb_termes=("term", "count"),
        clics=("clicks", "sum"),
        conversions=("conversions", "sum"),
    ).reset_index().sort_values("depense", ascending=False)
    agg.columns = ["Marque concurrente", "Dépense (€)", "Nb termes", "Clics", "Conversions"]
    return agg


def build_opportunities(df: pd.DataFrame) -> pd.DataFrame:
    sub = df[(df["decision"] == "keep") & (df["conversions"] > 0)].copy()
    status_lower = sub["status"].astype(str).str.lower()
    sub = sub[~status_lower.str.contains("ajouté|added", regex=True, na=False)]
    if sub.empty:
        return pd.DataFrame([{"Message": "Aucune opportunité détectée (tous les termes convertisseurs sont déjà des mots-clés actifs)."}])
    sub = sub.sort_values("conversions", ascending=False)
    out = sub[["term", "conversions", "conv_value", "cost", "clicks", "theme"]].copy()
    out["CPA (€)"] = out.apply(lambda r: (r["cost"] / r["conversions"]) if r["conversions"] else 0, axis=1)
    out = out[["term", "conversions", "conv_value", "cost", "CPA (€)", "clicks", "theme"]]
    out.columns = ["Terme à promouvoir", "Conversions", "Valeur conv. (€)",
                   "Coût (€)", "CPA (€)", "Clics", "Thème"]
    return out


def build_themes(df: pd.DataFrame) -> pd.DataFrame:
    agg = df.groupby(["theme", "decision"]).agg(
        nb_termes=("term", "count"),
        cost=("cost", "sum"),
        conversions=("conversions", "sum"),
    ).reset_index()
    pivot = agg.pivot_table(index="theme", columns="decision", values="nb_termes", fill_value=0)
    for col in ("keep", "exclude", "review"):
        if col not in pivot.columns:
            pivot[col] = 0
    theme_cost = df.groupby("theme")["cost"].sum()
    theme_conv = df.groupby("theme")["conversions"].sum()
    out = pd.DataFrame({
        "Thème": pivot.index,
        "Nb termes (total)": pivot.sum(axis=1).values,
        "À garder": pivot["keep"].values,
        "À exclure": pivot["exclude"].values,
        "À revoir": pivot["review"].values,
        "Dépense totale (€)": [theme_cost[t] for t in pivot.index],
        "Conversions": [theme_conv[t] for t in pivot.index],
    }).sort_values("Dépense totale (€)", ascending=False).reset_index(drop=True)
    return out


def write_xlsx(out_path: Path, tabs: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, data in tabs.items():
            data.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        for name, data in tabs.items():
            ws = wb[name[:31]]
            if ws.max_row > 1:
                style_header(ws, ws.max_column)
                autosize_columns(ws)
                if "Décision" in list(data.columns):
                    col_idx = list(data.columns).index("Décision") + 1
                    color_decision(ws, col_idx, len(data))


def load_objective(arg: str | None) -> str:
    if arg is None:
        print("Décris l'objectif de la campagne (2-3 phrases). Ex: « B2B consulting Google Ads 5k€+/mois, "
              "éviter formation gratuite et emploi ». Termine par une ligne vide:", file=sys.stderr)
        lines = []
        while True:
            try:
                line = input()
            except EOFError:
                break
            if not line.strip():
                break
            lines.append(line)
        return "\n".join(lines).strip()
    p = Path(arg)
    if p.exists() and p.is_file():
        return p.read_text(encoding="utf-8").strip()
    return arg.strip()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Audite un rapport de termes de recherche Google Ads et produit un xlsx d'analyse.",
    )
    parser.add_argument("csv", type=Path, help="Chemin vers le CSV exporté depuis Google Ads")
    parser.add_argument("-o", "--output", type=Path, default=None, help="Fichier xlsx de sortie")
    parser.add_argument("--objective", type=str, default=None,
                        help="Objectif de la campagne (texte direct ou chemin vers un .txt)")
    parser.add_argument("--model", type=str, default=MODEL, help=f"Modèle Claude (défaut: {MODEL})")
    parser.add_argument("--batch-size", type=int, default=BATCH_SIZE,
                        help=f"Nombre de termes par appel API (défaut: {BATCH_SIZE})")
    args = parser.parse_args()

    if not args.csv.exists():
        print(f"Fichier introuvable: {args.csv}", file=sys.stderr)
        return 1
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ANTHROPIC_API_KEY non définie dans l'environnement.", file=sys.stderr)
        return 1

    out_path = args.output or args.csv.with_name(f"{args.csv.stem}_audit.xlsx")
    objective = load_objective(args.objective)
    if not objective:
        print("Objectif de campagne requis.", file=sys.stderr)
        return 1

    print(f"→ Lecture du CSV: {args.csv}")
    df = load_csv(args.csv)
    print(f"  {len(df)} termes uniques détectés")
    if df.empty:
        print("Aucun terme à analyser.", file=sys.stderr)
        return 1

    client = Anthropic()
    print(f"→ Classification ({args.model}, batches de {args.batch_size})")
    df = classify_all(client, objective, df, args.model, args.batch_size)

    print("→ Génération du xlsx")
    tabs = {
        "Synthèse": build_synthesis(df),
        "Top dépensiers": build_top_spenders(df),
        "Candidats négatifs": build_candidates(df),
        "Concurrents détectés": build_competitors(df),
        "Opportunités": build_opportunities(df),
        "Thématiques": build_themes(df),
    }
    write_xlsx(out_path, tabs)
    print(f"✓ Rapport écrit: {out_path}")

    excl = int((df["decision"] == "exclude").sum())
    wasted = float(df.loc[df["decision"] == "exclude", "cost"].sum())
    print(f"  {excl} termes à exclure · {wasted:.2f} € d'économie potentielle")
    return 0


if __name__ == "__main__":
    sys.exit(main())
